# -*- coding: utf-8 -*-
import hashlib
import datetime, time
from flask import render_template, redirect, url_for, abort, flash, request,\
    current_app, make_response, send_file
from jinja2 import PackageLoader, Environment
from flask_login import login_required, current_user
from flask_sqlalchemy import Pagination
from flask_babelex import gettext, lazy_gettext
from io import BytesIO, StringIO
import xhtml2pdf.pisa as pisa
import flask_whooshalchemyplus
import barcode
from barcode.writer import ImageWriter
from openpyxl.reader.excel import load_workbook
from openpyxl.workbook import Workbook
import time
from decimal import Decimal
from . import main
from .. import db, uploader
from ..decorators import user_has
from ..utils import gen_serial_no, Master, full_response, custom_response, R201_CREATED, R400_BADREQUEST, R200_OK,\
    timestamp, datestr_to_timestamp
from ..constant import ORDER_EXCEL_FIELDS, HUAZHU_ORDER_STATUS, SORT_TYPE_CODE, ORDER_EXPRESS_FIELDS, MIXPUS_ORDER_FIELDS
from app.models import Product, Order, OrderItem, OrderStatus, Warehouse, Store, ProductStock, ProductSku, Express,\
    OutWarehouse, StockHistory, Site, Supplier, Asset, Shipper
from app.forms import OrderForm, OrderExpressForm, OrderRemark
from .filters import supress_none, timestamp2string, break_line
from app.helpers import kdniao

status_list = (
    (OrderStatus.PENDING_PAYMENT, lazy_gettext('Pending Payment')),
    (OrderStatus.PENDING_CHECK, lazy_gettext('Pending Check')),
    (OrderStatus.PENDING_SHIPMENT, lazy_gettext('Pending Shipment')),
    (OrderStatus.SHIPPED, lazy_gettext('Shipped')),
    (OrderStatus.SIGNED, lazy_gettext('Signed')),
    (OrderStatus.FINISHED, lazy_gettext('Pending Finished')),
    (OrderStatus.REFUND, lazy_gettext('Refund'))
)

def load_common_data():
    """
    私有方法，装载共用数据
    """
    pending_pay_count = Order.query.filter_by(master_uid=Master.master_uid(),status=OrderStatus.PENDING_PAYMENT).count()
    pending_review_count = Order.query.filter_by(master_uid=Master.master_uid(),status=OrderStatus.PENDING_CHECK).count()
    pending_ship_count = Order.query.filter_by(master_uid=Master.master_uid(),status=OrderStatus.PENDING_SHIPMENT).count()

    # 库房列表
    warehouse_list = Warehouse.query.filter_by(master_uid=Master.master_uid()).all()
    # 店铺列表
    store_list = Store.query.filter_by(master_uid=Master.master_uid(), status=1).all()

    return {
        'pending_pay_count': pending_pay_count,
        'pending_review_count': pending_review_count,
        'pending_ship_count': pending_ship_count,
        'warehouse_list': warehouse_list,
        'store_list': store_list,
        'top_menu': 'orders',
        'status_list': status_list
    }


@main.route('/orders')
@main.route('/orders/<int:page>')
@login_required
@user_has('admin_order')
def show_orders(page=1):
    per_page = request.args.get('per_page', 10, type=int)
    status = request.args.get('s', 0, type=int)

    query = Order.query.filter_by(master_uid=Master.master_uid())
    if status:
        query = query.filter_by(status=status)

    paginated_orders = query.order_by('created_at desc').paginate(page, per_page)

    return render_template('orders/show_list.html',
                           sub_menu='orders',
                           status=status,
                           paginated_orders=paginated_orders.items,
                           pagination=paginated_orders,
                           **load_common_data())


@main.route('/orders/search', methods=['GET', 'POST'])
@login_required
@user_has('admin_order')
def search_orders(page=1):
    """订单搜索"""
    per_page = request.values.get('per_page', 10, type=int)
    page = request.values.get('page', 1, type=int)
    qk = request.values.get('qk')
    store_id = request.values.get('store_id', type=int)
    s = request.values.get('s', type=int)
    date = request.values.get('date', type=int)
    sk = request.values.get('sk', type=str, default='ad')

    builder = Order.query.filter_by(master_uid=Master.master_uid())

    qk = qk.strip()
    if qk:
        current_app.logger.warn('Search order [%s]!' % qk)
        builder = builder.whoosh_search(qk, or_=True)
    
    if store_id:
        builder = builder.filter_by(store_id=store_id)
    if s:
        builder = builder.filter_by(status=s)
        
    if date:
        now = datetime.date.today()
        if date == 1: # 今天之内
            start_time = time.mktime((now.year, now.month, now.day, 0, 0, 0, 0, 0, 0))
            end_time = time.time()
        if date == 2: # 昨天之内
            dt = now - datetime.timedelta(days=1)
            start_time = time.mktime((dt.year, dt.month, dt.day, 0, 0, 0, 0, 0, 0))
            end_time = time.mktime((now.year, now.month, now.day, 0, 0, 0, 0, 0, 0))
        if date == 7: # 7天之内
            dt = now - datetime.timedelta(days=7)
            start_time = time.mktime((dt.year, dt.month, dt.day, 0, 0, 0, 0, 0, 0))
            end_time = time.mktime((now.year, now.month, now.day, 0, 0, 0, 0, 0, 0))
        if date == 30: # 30天之内
            dt = now - datetime.timedelta(days=30)
            start_time = time.mktime((dt.year, dt.month, dt.day, 0, 0, 0, 0, 0, 0))
            end_time = time.mktime((now.year, now.month, now.day, 0, 0, 0, 0, 0, 0))

        builder = builder.filter(Order.created_at > start_time, Order.created_at < end_time)

    orders = builder.order_by(Order.created_at.desc()).all()

    # 构造分页
    total_count = builder.count()
    if page == 1:
        start = 0
    else:
        start = (page - 1) * per_page
    end = start + per_page

    current_app.logger.debug('total count [%d], start [%d], per_page [%d]' % (total_count, start, per_page))

    paginated_orders = orders[start:end]

    pagination = Pagination(query=None, page=page, per_page=per_page, total=total_count, items=None)

    return render_template('orders/search_result.html',
                           qk=qk,
                           sk=sk,
                           store_id=store_id,
                           s=s,
                           date=date,
                           paginated_orders=paginated_orders,
                           pagination=pagination)

@main.route('/orders/<string:rid>/ajax_add_express', methods=['GET', 'POST'])
@login_required
@user_has('admin_order')
def order_express_no(rid):
    """手动设置物流单号，开始发货"""
    order = Order.query.filter_by(master_uid=Master.master_uid(), serial_no=rid).first()
    if order is None:
        return custom_response(False, gettext("Order isn't exist!")) if request.method == 'POST' else abort(404)

    form = OrderExpressForm()
    if form.validate_on_submit():
        # 验证是否有效
        if order.status in [OrderStatus.PENDING_PAYMENT, OrderStatus.PENDING_CHECK]:
            return custom_response(False, gettext('Order status is error!'))

        # 点击发货
        order.express_id = form.express_id.data
        order.express_no = form.express_no.data
        order.mark_shipped_status()

        # 自动生成出库单
        out_serial_no = gen_serial_no('CK')
        out_warehouse = OutWarehouse(
            master_uid = Master.master_uid(),
            serial_no = out_serial_no,
            target_serial_no = order.serial_no,
            target_type = 1,
            warehouse_id = order.warehouse_id,
            total_quantity = order.total_quantity,
            out_quantity = 0,
            # 出库状态： 1、未出库 2、出库中 3、出库完成
            out_status = 1,
            # 出库流程状态
            status = 1
        )
        db.session.add(out_warehouse)

        # 获取库房信息
        warehouse = Warehouse.query.get(order.warehouse_id)
        if not warehouse:
            return custom_response(False, gettext("Warehouse info isn't exist!"))
        default_shelve = warehouse.default_shelve

        # 出库单明细
        for item in order.items:
            sku_id = item.sku_id
            product_stock = ProductStock.query.filter_by(product_sku_id=sku_id,
                                                         warehouse_id=order.warehouse_id).first()
            if product_stock is None:
                return custom_response(False, gettext('This item is out of stock'))

            offset_quantity = item.quantity
            current_quantity = product_stock.current_count
            original_quantity = current_quantity + offset_quantity

            stock_history = StockHistory(
                master_uid = Master.master_uid(),
                warehouse_id = order.warehouse_id,
                warehouse_shelve_id = default_shelve.id,
                product_sku_id = sku_id,
                sku_serial_no = item.sku_serial_no,

                # 出库单/入库单 编号
                serial_no = out_serial_no,
                # 类型：1、入库 2：出库
                type = 2,
                # 操作类型
                operation_type = 21,
                # 原库存数量
                original_quantity = original_quantity,
                # 变化数量
                quantity = offset_quantity,
                # 当前数量
                current_quantity = current_quantity,
                # 原价格
                ori_price = item.deal_price,
                price = item.deal_price
            )

            db.session.add(stock_history)

        db.session.commit()

        return custom_response(True, 'Add express info if ok!')

    form.express_id.data = order.express_id
    form.express_no.data = order.express_no

    # get express
    express_list = Express.query.filter_by(master_uid=Master.master_uid()).all()

    return render_template('orders/_modal_express.html',
                           order=order,
                           form=form,
                           express_list=express_list,
                           post_url=url_for('.order_express_no', rid=rid))


@main.route('/orders/shipment', methods=['POST'])
@login_required
@user_has('admin_order')
def shipment_order():
    """获取订单的电子面单"""
    rid = request.values.get('rid')
    rids = rid.split(',')
    order_list = Order.query.filter_by(master_uid=Master.master_uid()).filter(Order.serial_no.in_(rids)).all()

    messages = {}
    for order in order_list:
        cur_rid = order.rid
        # 检查订单是否为待发货
        if order.status != OrderStatus.PENDING_SHIPMENT:
            current_app.logger.warn("Order[%s] status isn't pending shipment!" % order.rid)
            messages[cur_rid] = gettext("Order[%s] status isn't pending shipment!" % order.rid)
            continue

        # 获取客户设置快递公司，如无则选择默认
        express_id = order.express_id

        express = None
        if express_id:
            express = Express.query.get(express_id)

        # 获取默认值
        if express is None:
            express = Express.query.filter_by(master_uid=Master.master_uid(), is_default=True).first()

        if express is None:
            current_app.logger.warn("Order[%s] express not set!" % order.rid)
            messages[cur_rid] = gettext("Order[%s] express not set!" % order.rid)
            continue

        # 发货前，需先审单，设置从哪个仓库发出
        warehouse_id = order.warehouse_id
        # 获取发货人
        shipper = Shipper.query.filter_by(master_uid=Master.master_uid(), warehouse_id=warehouse_id).first()

        eorder = {}
        eorder['ShipperCode'] = express.code
        # 同一电子面单模板可多次调用，无需保存到本地
        eorder['LogisticCode'] = order.express_no
        eorder['OrderCode'] = order.outside_target_id if order.outside_target_id else order.rid
        eorder['PayType'] = 1
        eorder['ExpType'] = 1
        # 返回电子面单模板：0-不需要；1-需要
        eorder['IsReturnPrintTemplate'] = '0'

        # 电子面单客户账号（与快递网点申请）
        eorder['CustomerName'] = express.customer_name
        eorder['CustomerPwd'] = express.customer_pwd
        eorder['SendSite'] = express.send_site

        sender = {}
        sender['Name'] = shipper.name
        sender['Mobile'] = shipper.mobile
        sender['ProvinceName'] = shipper.province
        sender['CityName'] = shipper.city
        sender['ExpAreaName'] = shipper.area
        sender['Address'] = shipper.address

        receiver = {}
        receiver['Name'] = order.buyer_name
        receiver['Mobile'] = order.buyer_phone
        receiver['ProvinceName'] = order.buyer_province
        receiver['CityName'] = order.buyer_city
        receiver['ExpAreaName'] = ''
        receiver['Address'] = order.buyer_address

        commodity_one = {}
        commodity_one['GoodsName'] = '其他'
        commodity = []
        commodity.append(commodity_one)

        eorder['Sender'] = sender
        eorder['Receiver'] = receiver
        eorder['Commodity'] = commodity

        # 请求电子面单
        eorder_result = kdniao.get_eorder(eorder)

        current_app.logger.debug('result code: %s, success: %s' % (eorder_result['ResultCode'], eorder_result['Success']))

        if eorder_result['ResultCode'] != '100' or eorder_result['Success'] != True:
            reason = eorder_result['Reason']
            messages[cur_rid] = reason
            continue

        # 获取电子面单成功，则填充物流单号
        logistic_code = eorder_result['Order']['LogisticCode']

        # 点击发货
        order.express_no = logistic_code
        order.mark_print_status()


    db.session.commit()

    success = False if len(messages) else True

    return custom_response(success=success, message=messages)


@main.route('/orders/print_eorder', methods=['POST'])
@login_required
@user_has('admin_order')
def print_eorder():
    """打印电子面单"""
    rid = request.values.get('rid')
    current_order = Order.query.filter_by(master_uid=Master.master_uid(), serial_no=rid).first()

    # 检查订单是否为待发货
    if current_order.status != OrderStatus.PENDING_PRINT or not current_order.express_no:
        current_app.logger.warn("Order[%s] status isn't pending print!" % rid)
        return custom_response(False, gettext("Order[%s] status is error or express_no is empty!" % rid))

    # 获取客户设置快递公司，如无则选择默认
    express_id = current_order.express_id

    express = None
    if express_id:
        express = Express.query.get(express_id)

    # 获取默认值
    if express is None:
        express = Express.query.filter_by(master_uid=Master.master_uid(), is_default=True).first()

    # 获取发货人
    shipper = Shipper.query.filter_by(master_uid=Master.master_uid(), warehouse_id=current_order.warehouse_id).first()

    eorder = {}
    eorder['ShipperCode'] = express.code
    # 同一电子面单模板可多次调用，无需保存到本地
    eorder['LogisticCode'] = current_order.express_no
    eorder['OrderCode'] = current_order.outside_target_id if current_order.outside_target_id else current_order.rid
    eorder['PayType'] = 1
    eorder['ExpType'] = 1
    eorder['IsReturnPrintTemplate'] = '1'

    eorder['CustomerName'] = express.customer_name
    eorder['CustomerPwd'] = express.customer_pwd
    eorder['SendSite'] = express.send_site

    sender = {}
    sender['Name'] = shipper.name
    sender['Mobile'] = shipper.mobile
    sender['ProvinceName'] = shipper.province
    sender['CityName'] = shipper.city
    sender['ExpAreaName'] = shipper.area
    sender['Address'] = shipper.address

    receiver = {}
    receiver['Name'] = current_order.buyer_name
    receiver['Mobile'] = current_order.buyer_phone
    receiver['ProvinceName'] = current_order.buyer_province
    receiver['CityName'] = current_order.buyer_city
    receiver['ExpAreaName'] = ''
    receiver['Address'] = current_order.buyer_address

    commodity_one = {}
    commodity_one['GoodsName'] = '其他'
    commodity = []
    commodity.append(commodity_one)

    eorder['Sender'] = sender
    eorder['Receiver'] = receiver
    eorder['Commodity'] = commodity

    # 请求电子面单
    eorder_result = kdniao.get_eorder(eorder)

    current_app.logger.debug(
        'result code: %s, success: %s' % (eorder_result['ResultCode'], eorder_result['Success']))

    current_app.logger.debug('%s' % eorder_result)

    if eorder_result['ResultCode'] == '105' and eorder_result['Success'] == False:
        return full_response(data=eorder_result['PrintTemplate'])

    if eorder_result['ResultCode'] != '100' or eorder_result['Success'] != True:
        reason = eorder_result['Reason']
        return custom_response(False, reason)


@main.route('/orders/print_order_pdf')
@login_required
@user_has('admin_order')
def print_order_pdf():
    """打印发货单"""
    rid = request.args.get('rid')
    rids = rid.split(',')
    preview = request.args.get('preview')
    order_list = Order.query.filter_by(master_uid=Master.master_uid()).filter(Order.serial_no.in_(rids)).all()

    # 同步自动生成出库单
    auto_gen_outwarehouse(order_list)

    env = Environment(loader=PackageLoader(current_app.name, 'templates'))
    env.filters['supress_none'] = supress_none
    env.filters['timestamp2string'] = timestamp2string
    env.filters['break_line'] = break_line
    template = env.get_template('pdf/order.html')

    current_site = Site.query.filter_by(master_uid=Master.master_uid()).first()

    title_attrs = {
        'bill_name': gettext('Shipping Bill'),
        'store_name': gettext('Store Name'),
        'serial_no': gettext('Order Serial'),
        'consignee': gettext('Consignee'),
        'date': gettext('Payed Date'),
        'consignee_name': gettext('Consignee name'),
        'remark': gettext('Buyer Remark'),
        'product_items': gettext('Product Items'),

        'order_number': gettext('Order Number'),
        'sn': gettext('Product Serial'),
        'product_name': gettext('Product Name'),
        'mode': gettext('Product Mode'),
        'unit': gettext('Unit'),
        'quantity': gettext('Quantity'),
        'price': gettext('Price'),
        'discount_price': gettext('Discount Price'),
        'subtotal': gettext('Subtotal'),
        'total': gettext('Total'),

        'freight': gettext('Freight'),
        'discount': gettext('Discount'),
        'pay_amount': gettext('Pay Amount')
    }

    font_path = 'http://s3.mixpus.com/static/fonts/simsun.ttf'
    if current_app.config['MODE'] == 'dev':
        font_path = current_app.root_path + '/static/fonts/simsun.ttf'

    html = template.render(
        current_site=current_site,
        title_attrs=title_attrs,
        font_path=font_path,
        order_list=order_list,
    ).encode('utf-8')

    if preview:
        return html

    result = BytesIO()
    pdf = pisa.CreatePDF(BytesIO(html), result)
    resp = make_response(result.getvalue())
    export_file = 'Order-{}'.format(int(timestamp()))
    resp.headers['Content-Disposition'] = ("inline; filename='{0}'; filename*=UTF-8''{0}".format(export_file))
    resp.headers['Content-Type'] = 'application/pdf'

    return resp


def auto_gen_outwarehouse(order_list):
    """同步自动生成出库单"""

    for order in order_list:
        # 检测是否已存在出库单
        if OutWarehouse.query.filter_by(master_uid=Master.master_uid(),
                                        target_serial_no=order.serial_no,warehouse_id=order.warehouse_id).first():
            # 已存在，则跳过
            continue

        out_serial_no = gen_serial_no('CK')
        out_warehouse = OutWarehouse(
            master_uid=Master.master_uid(),
            serial_no=out_serial_no,
            target_serial_no=order.serial_no,
            target_type=1,
            warehouse_id=order.warehouse_id,
            total_quantity=order.total_quantity,
            out_quantity=0,
            # 出库状态： 1、未出库 2、出库中 3、出库完成
            out_status=1,
            # 出库流程状态
            status=1
        )
        db.session.add(out_warehouse)

        # 获取库房信息
        warehouse = Warehouse.query.get(order.warehouse_id)
        if not warehouse:
            return custom_response(False, gettext("Warehouse info isn't exist!"))
        default_shelve = warehouse.default_shelve

        # 出库单明细
        for item in order.items:
            sku_id = item.sku_id
            product_stock = ProductStock.query.filter_by(product_sku_id=sku_id,
                                                         warehouse_id=order.warehouse_id).first()
            if product_stock is None:
                return custom_response(False, gettext('This item is out of stock'))

            offset_quantity = item.quantity
            current_quantity = product_stock.current_count
            original_quantity = current_quantity + offset_quantity

            stock_history = StockHistory(
                master_uid=Master.master_uid(),
                warehouse_id=order.warehouse_id,
                warehouse_shelve_id=default_shelve.id,
                product_sku_id=sku_id,
                sku_serial_no=item.sku_serial_no,

                # 出库单/入库单 编号
                serial_no=out_serial_no,
                # 类型：1、入库 2：出库
                type=2,
                # 操作类型
                operation_type=21,
                # 原库存数量
                original_quantity=original_quantity,
                # 变化数量
                quantity=offset_quantity,
                # 当前数量
                current_quantity=current_quantity,
                # 原价格
                ori_price=item.deal_price,
                price=item.deal_price
            )

            db.session.add(stock_history)

    db.session.commit()



def get_key_by_value(dict_value):
    """通过值获取key"""
    for (key, value) in ORDER_EXCEL_FIELDS.items():
        if value == dict_value:
            return key
    return None


def get_status_by_value(val):
    """通过值获取订单状态"""
    for t in HUAZHU_ORDER_STATUS:
        if t[1] == val:
            return t[0]


def import_order_by_dict(order_list, store_id, warehouse_id):
    """导入订单数据"""

    bad_orders = []
    bad_skus = []

    # 默认物流公司未设置
    default_express = Express.query.filter_by(master_uid=Master.master_uid(), is_default=True).first()
    if default_express is None:
        flash(gettext("Default express isn't setting!"), 'danger')
        return render_template('orders/import_result.html',
                               bad_skus=bad_skus,
                               bad_orders=bad_orders)

    for order_info in order_list:
        # 验证是否已经存在
        if Order.query.filter_by(master_uid=Master.master_uid(), outside_target_id=order_info['order_serial_no']).first():
            current_app.logger.warn("Order[%s] is already imported!" % order_info['order_serial_no'])
            bad_orders.append(order_info['order_serial_no'])
            continue

        # 获取订单里产品
        product = {
            'name': order_info['order_product_name'],
            's_model': order_info['s_model'],
            'quantity': order_info['quantity'],
            'cost_price': order_info['cost_price'],
            'sale_price': order_info['sale_price'],
            'sku_serial_no': order_info['store_product_id'].strip(), # 截取前后空格
            'discount_total_amount': order_info['discount_total_amount']
        }

        # 验证产品是否存在
        current_sku = ProductSku.query.filter_by(master_uid=Master.master_uid(), serial_no=product['sku_serial_no']).first()

        if current_sku is None:
            # 产品不存在
            current_app.logger.warn("Current sku[%s] isn't exist!" % product['sku_serial_no'])
            bad_skus.append((product['sku_serial_no'], 1))
            continue
        
        # 开始导入订单
        new_order_serial_no = Order.make_unique_serial_no(gen_serial_no('C'))

        current_app.logger.debug('Ordered_at: %s' % order_info['ordered_at'])

        ordered_at = order_info['ordered_at']
        if ordered_at[:-2] == '.0':
            ordered_at = ordered_at[:-2]
            ordered_at = time.mktime(time.strptime(order_info['ordered_at'], '%Y-%m-%d %H:%M:%S'))
        else:
            ordered_at = timestamp()

        order = Order(
            master_uid = Master.master_uid(),
            outside_target_id = order_info['order_serial_no'],
            serial_no = new_order_serial_no,
            store_id = store_id,
            warehouse_id = warehouse_id,

            pay_amount = order_info['total_amount'] + order_info['freight'] - order_info['discount_total_amount'],
            total_quantity = order_info['quantity'],
            total_amount = order_info['total_amount'],
            freight = order_info['freight'],
            discount_amount = order_info['discount_total_amount'],

            express_id = default_express.id,
            express_no = order_info['express_no'],
            status=get_status_by_value(order_info['order_status']),
            remark = '',
            # 顾客信息
            buyer_name = order_info['buyer_name'],
            buyer_tel = order_info['buyer_phone'],
            buyer_phone = order_info['buyer_mobile'],
            buyer_address = order_info['buyer_address'],
            buyer_zipcode = '',
            buyer_country = '中国',
            buyer_province = '',
            buyer_city = '',
            buyer_remark = order_info['buyer_remark'],
            created_at = ordered_at,
            updated_at = ordered_at
        )

        db.session.add(order)

        order_item = OrderItem(
            order = order,
            order_serial_no = new_order_serial_no,
            sku_id = current_sku.id,
            sku_serial_no = current_sku.serial_no,
            quantity = product['quantity'],
            deal_price = product['sale_price'],
            discount_amount = product['discount_total_amount']
        )

        db.session.add(order_item)

    # 新增索引
    flask_whooshalchemyplus.index_one_model(Order)
    
    db.session.commit()
    
    if bad_orders or bad_skus:
        flash(gettext('Import orders is error!!!'), 'danger')
    else:
        flash(gettext("Import Orders is ok!"), 'success')

    return render_template('orders/import_result.html',
                           bad_skus=bad_skus,
                           bad_orders=bad_orders)


@main.route('/orders/import', methods=['GET', 'POST'])
@user_has('admin_order')
@login_required
def import_orders():
    """导入订单"""
    if request.method == 'POST':
        wh_id = request.form.get('wh_id', type=int)
        st_id = request.form.get('store_id', type=int)
        f = request.files['excel']

        # start to save
        sub_folder = str(time.strftime('%y%m%d'))
        name_prefix = 'mix' + str(time.time())
        name_prefix = hashlib.md5(name_prefix.encode('utf-8')).hexdigest()[:15]
        filename = uploader.save(f, folder=sub_folder, name=name_prefix + '.')

        order_file = uploader.path(filename)

        current_app.logger.debug('Excel file [%s]' % order_file)

        # 读取文件
        wb = load_workbook(filename=order_file)
        sheets = wb.get_sheet_names()
        # 默认第一个表格的名称
        ws = wb.get_sheet_by_name(sheets[0])

        # 获取表格所有行和列，两者都是可迭代的
        # rows = ws.rows
        # columns = ws.columns

        # 获取行数
        total_rows = ws.max_row
        # 获取列数
        total_cols = ws.max_column

        # 获取表头
        counter = 1
        header = []
        for row in ws.rows:
            if counter == 1: # 仅获取第一行
                header = [col.value for col in row]
            counter += 1

        # 文件转化为列表
        order_list = []
        for row_idx in range(2, total_rows + 1):
            order_info = {}
            for col_idx in range(1, total_cols + 1):
                key = get_key_by_value(header[col_idx - 1])
                cell_value = ws.cell(row=row_idx, column=col_idx).value

                current_app.logger.debug('%s:%s' % (key, cell_value))

                order_info[key] = cell_value

            order_list.append(order_info)

        # 开始导入订单
        return import_order_by_dict(order_list, st_id, wh_id)

    # 库房列表
    warehouse_list = Warehouse.query.filter_by(master_uid=Master.master_uid()).all()
    # 店铺列表
    store_list = Store.query.filter_by(master_uid=Master.master_uid(), status=1).all()

    return render_template('orders/_modal_import.html',
                           store_list=store_list,
                           warehouse_list=warehouse_list)


def _rebuild_header_value(key, current_site):
    """重建Excel表头标题"""
    if key in ['freight','extra_charge','total_amount',]:
        return '{}({})'.format(ORDER_EXCEL_FIELDS[key], current_site.currency)
    return ORDER_EXCEL_FIELDS[key]


@main.route('/orders/export', methods=['GET', 'POST'])
@login_required
@user_has('admin_order')
def export_orders():
    """导出订单"""
    if request.method == 'POST':
        store_id = request.form.get('store_id', type=int)
        status = request.form.getlist('status[]')
        start_date = request.form.get('start_date')
        end_date = request.form.get('end_date')
        type = request.form.get('type', 1, type=int)

        builder = Order.query.filter_by(master_uid=Master.master_uid())
        if store_id:
            builder = builder.filter_by(store_id=store_id)

        if status:
            status = [int(s) for s in status]
            builder = builder.filter(Order.status.in_(status))

        if start_date:
            start_date = datestr_to_timestamp(start_date)
            builder = builder.filter(Order.created_at >= start_date)

        if end_date:
            end_date = datestr_to_timestamp(end_date)
            builder = builder.filter(Order.created_at <= end_date)

        order_list = builder.order_by('created_at asc').all()

        filename = 'mic_order_list'
        if type == 1:
            filename = 'mic_order_express'

        dest_filename = r'{}_{}.xlsx'.format(filename, datetime.datetime.now().strftime('%Y%m%d'))
        export_path = current_app.root_path + '/static/'
        export_file = '{}{}'.format(export_path, dest_filename)

        # 新建文件
        wb = Workbook()
        # 第一个sheet
        ws = wb.active
        ws.title = 'Order List'

        current_app.logger.debug('Export order count: %d' % len(order_list))

        # 导出订单物流信息
        if type == 1:
            express_columns = [key for key in ORDER_EXPRESS_FIELDS.keys()]
            # 写入表头
            for col in range(1, len(express_columns) + 1):
                field = express_columns[col - 1]
                ws.cell(row=1, column=col, value=ORDER_EXPRESS_FIELDS[field])

            # 写入数据，从第2行开始写入
            for row in range(2, len(order_list) + 2):
                current_order = order_list[row - 2]
                for col in range(1, len(express_columns) + 1):
                    field = express_columns[col - 1]

                    # outside_target_id, express_no
                    cell_value = current_order.to_json().get(field)

                    ws.cell(row=row, column=col, value=cell_value)
        # 导出订单信息
        else:
            order_columns = [key for key in MIXPUS_ORDER_FIELDS.keys()]

            # 写入表头
            for col in range(1, len(order_columns) + 1):
                field = order_columns[col - 1]
                ws.cell(row=1, column=col, value=MIXPUS_ORDER_FIELDS[field])

            # 写入数据, 从第2行开始写入
            for row in range(2, len(order_list) + 2):
                current_order = order_list[row - 2]

                current_app.logger.debug('Current row: %d' % row)

                items = current_order.items

                for col in range(1, len(order_columns) + 1):
                    field = order_columns[col - 1]

                    cell_value = current_order.to_json().get(field)
                    if field == 'status':
                        cell_value = str(cell_value)
                    elif field in ['created_at', 'express_at', 'received_at']:
                        cell_value = timestamp2string(cell_value) if cell_value else ''
                    elif field == 'product_id':
                        cell_value = ';'.join([item.sku_serial_no for item in items])
                    elif field == 'product_name':
                        cell_value = ';'.join([item.sku.product_name for item in items])
                    elif field == 's_model':
                        cell_value = ';'.join(['{}{}'.format(item.sku.s_model, item.sku.s_color) for item in items])
                    elif field == 'deal_price':
                        cell_value = ';'.join([str(item.deal_price) for item in items])
                    else:
                        current_app.logger.debug('Current col: %s' % field)

                    ws.cell(row=row, column=col, value=cell_value)

        wb.save(filename=export_file)

        return full_response(True, R200_OK, {'download_url': url_for('static', filename=dest_filename)})

    # 店铺列表
    store_list = Store.query.filter_by(master_uid=Master.master_uid(), status=1).all()

    return render_template('orders/_modal_export.html',
                           post_url=url_for('main.export_orders'),
                           store_list=store_list,
                           status_list=status_list)


@main.route('/orders/create', methods=['GET', 'POST'])
@login_required
@user_has('admin_order')
def create_order():
    form = OrderForm()
    # 获取店铺
    store_list = Store.query.filter_by(master_uid=Master.master_uid(), status=1).all()
    # 获取仓库列表
    warehouse_list = Warehouse.query.filter_by(master_uid=Master.master_uid(), status=1).all()
    # 获取快递类型
    express_list = Express.query.filter_by(master_uid=Master.master_uid()).all()

    # 初始化选项
    form.store_id.choices = [(store.id, '%s -- %s' % (store.name, store.platform_name)) for store in store_list]
    form.warehouse_id.choices = [(warehouse.id, warehouse.name) for warehouse in warehouse_list]
    form.express_id.choices = [(express.id, express.name) for express in express_list]

    if request.method == 'POST':
        # 验证必须字段
        if not form.validate_on_submit():
            current_app.logger.debug(form.errors)
            return full_response(False, R400_BADREQUEST, form.errors)

        items = request.form.getlist('items[]')
        if not items or items is None:
            return custom_response(False, 'Order items is Null!')

        total_quantity = 0
        total_amount = 0
        total_discount = 0
        order_items = []
        for sku_id in items:
            sku_id = int(sku_id)
            quantity = request.form.get('quantity[%d]' % sku_id, type=int)
            discount = request.form.get('discount[%d]' % sku_id, type=float)

            # 验证sku信息
            product_sku = ProductSku.query.get(sku_id)
            if not product_sku:
                return custom_response(False, gettext('Product sku[%d] is not exist!' % sku_id))

            # 验证库存
            product_stock = ProductStock.query.filter_by(product_sku_id=sku_id, warehouse_id=form.warehouse_id.data).first()
            if product_stock.available_count < quantity:
                return custom_response(False, gettext('[%s] Inventory is not enough!' % product_sku.serial_no))
            
            # 添加订单明细
            item = {
                'sku_id': sku_id,
                'sku_serial_no': product_sku.serial_no,
                'quantity': quantity,
                'deal_price': product_sku.sale_price, # 交易价格
                'discount_amount': Decimal(discount)  # 优惠金额
            }
            order_items.append(item)

            total_amount += product_sku.sale_price * quantity
            total_discount += discount
            total_quantity += quantity

        # 添加订单
        freight = form.freight.data
        pay_amount = Decimal(total_amount) + freight - Decimal(total_discount)
        new_serial_no = Order.make_unique_serial_no(form.serial_no.data)
        order = Order(
            master_uid=Master.master_uid(),
            serial_no=new_serial_no,
            outside_target_id=form.outside_target_id.data,
            store_id=form.store_id.data,
            warehouse_id=form.warehouse_id.data,
            pay_amount=Decimal(pay_amount),
            # 总金额
            total_amount=Decimal(total_amount),
            # 总数量
            total_quantity=total_quantity,
            discount_amount=Decimal(total_discount),
            freight=form.freight.data,
            express_id=form.express_id.data,
            remark=form.remark.data,
            buyer_name=form.buyer_name.data,
            buyer_tel=form.buyer_tel.data,
            buyer_phone=form.buyer_phone.data,
            buyer_zipcode=form.buyer_zipcode.data,
            buyer_address=form.buyer_address.data,
            buyer_country=form.buyer_country.data,
            buyer_province=form.buyer_province.data,
            buyer_city=form.buyer_city.data,
            # 买家备注
            buyer_remark=form.buyer_remark.data,
        )

        db.session.add(order)

        # 更新订单明细
        for item in order_items:
            item['order_serial_no'] = new_serial_no
            order_item = OrderItem(order=order, **item)
            db.session.add(order_item)

        # 新增索引
        flask_whooshalchemyplus.index_one_model(Order)

        db.session.commit()

        return full_response(True, R201_CREATED, {'next_url': url_for('.show_orders')})

    # 新增订单
    mode = 'create'
    # 添加默认订单号
    form.serial_no.data = gen_serial_no('C')

    return render_template('orders/create_and_edit.html',
                           form=form,
                           mode=mode,
                           current_order=None,
                           **load_common_data())


@main.route('/orders/<string:sn>/edit', methods=['GET', 'POST'])
@login_required
@user_has('admin_order')
def edit_order(sn):
    current_order = Order.query.filter_by(master_uid=Master.master_uid(), serial_no=sn).first()
    if current_order is None:
        abort(404)

    form = OrderForm()

    # 获取店铺
    store_list = Store.query.filter_by(master_uid=Master.master_uid(), status=1).all()
    # 获取仓库列表
    warehouse_list = Warehouse.query.filter_by(master_uid=Master.master_uid(), status=1).all()
    # 获取快递类型
    express_list = Express.query.filter_by(master_uid=Master.master_uid()).all()

    # 初始化选项
    form.store_id.choices = [(store.id, store.name) for store in store_list]
    form.warehouse_id.choices = [(warehouse.id, warehouse.name) for warehouse in warehouse_list]
    form.express_id.choices = [(express.id, express.name) for express in express_list]

    if request.method == 'POST':
        pass


    mode = 'edit'
    form.serial_no.data = current_order.serial_no
    form.outside_target_id.data = current_order.outside_target_id
    form.store_id.data = current_order.store_id
    form.express_id.data = current_order.express_id
    form.warehouse_id.data = current_order.warehouse_id
    form.freight.data = current_order.freight
    form.remark.data = current_order.remark

    form.buyer_name.data = current_order.buyer_name
    form.buyer_tel.data = current_order.buyer_tel
    form.buyer_phone.data = current_order.buyer_phone
    form.buyer_zipcode.data = current_order.buyer_zipcode
    form.buyer_address.data = current_order.buyer_address
    form.buyer_country.data = current_order.buyer_country
    form.buyer_province.data = current_order.buyer_province
    form.buyer_city.data = current_order.buyer_city
    form.buyer_remark.data = current_order.buyer_remark

    return render_template('orders/create_and_edit.html',
                           form=form,
                           mode=mode,
                           current_order=current_order,
                           **load_common_data())


@main.route('/orders/<string:sn>/add_remark', methods=['GET', 'POST'])
@login_required
@user_has('admin_order')
def add_remark(sn):
    form = OrderRemark()
    order = Order.query.filter_by(master_uid=Master.master_uid(), serial_no=sn).first()
    if form.validate_on_submit():
        order.remark = form.remark.data

        db.session.commit()

        return custom_response(True, gettext('Add Remark is ok!'))

    form.remark.data = order.remark
    return render_template('orders/_modal_remark.html',
                           form=form,
                           current_order=order,
                           **load_common_data())


@main.route('/orders/<string:sn>/show')
@login_required
@user_has('admin_order')
def preview_order(sn):
    order = Order.query.filter_by(master_uid=Master.master_uid(), serial_no=sn).first()

    return render_template('orders/preview_order.html',
                           current_order=order, **load_common_data())


@main.route('/orders/<string:sn>/split', methods=['GET', 'POST'])
@login_required
@user_has('admin_order')
def split_order(sn):
    current_order = Order.query.filter_by(master_uid=Master.master_uid(), serial_no=sn).first()
    if current_order is None:
        abort(404)

    if request.method == 'POST':
        selected_item_ids = request.form.getlist('selected[]')
        if not selected_item_ids or selected_item_ids is None:
            return custom_response(False, 'Split order items is NULL!')

        total_quantity = 0
        total_amount = 0
        total_discount = 0
        selected_items = []
        for item_id in selected_item_ids:
            item = OrderItem.query.get(item_id)
            if item:
                total_quantity += item.quantity
                total_amount += item.quantity*item.deal_price
                total_discount += item.discount_amount

                selected_items.append(item)

        # 添加订单
        freight = 0
        pay_amount = Decimal(total_amount) + freight - Decimal(total_discount)
        new_serial_no = Order.make_unique_serial_no(gen_serial_no('C'))
        new_order = Order(
            master_uid=current_order.master_uid,
            serial_no=new_serial_no,
            outside_target_id=current_order.serial_no,
            store_id=current_order.store_id,
            warehouse_id=current_order.warehouse_id,
            pay_amount=Decimal(pay_amount),
            # 总金额
            total_amount=Decimal(total_amount),
            # 总数量
            total_quantity=total_quantity,
            discount_amount=Decimal(total_discount),
            freight=freight,
            express_id=current_order.express_id,
            remark=current_order.remark,
            buyer_name=current_order.buyer_name,
            buyer_tel=current_order.buyer_tel,
            buyer_phone=current_order.buyer_phone,
            buyer_zipcode=current_order.buyer_zipcode,
            buyer_address=current_order.buyer_address,
            buyer_country=current_order.buyer_country,
            buyer_province=current_order.buyer_province,
            buyer_city=current_order.buyer_city,
            # 买家备注
            buyer_remark=current_order.buyer_remark,
            type=2
        )
        db.session.add(new_order)

        # 更新订单明细
        for item in selected_items:
            item.order = new_order
            item.serial_no = new_serial_no

        # 更新旧订单
        current_order.pay_amount -= pay_amount
        current_order.total_amount -= total_amount
        current_order.total_quantity -= total_quantity
        current_order.discount_amount -= total_discount

        db.session.commit()

        return custom_response(True, gettext('Split order id ok!'), 200)

    return render_template('orders/_modal_split.html',
                           current_order=current_order,
                           **load_common_data())


@main.route('/orders/ajax_verify', methods=['POST'])
@login_required
@user_has('admin_order')
def ajax_verify_order():
    """订单审核"""
    selected_sns = request.form.getlist('selected[]')
    if not selected_sns or selected_sns is None:
        return custom_response(False, 'Verify order is NULL!')

    try:
        for sn in selected_sns:
            order = Order.query.filter_by(master_uid=Master.master_uid(), serial_no=sn).one()
            if order is None:
                return custom_response(False, gettext("Verify order isn't exist!"))
            
            if order.status == OrderStatus.PENDING_PAYMENT:
                # 已付款，进入待审核
                order.mark_checked_status()
            elif order.status == OrderStatus.PENDING_CHECK:
                # 完成审核，待发货状态

                # 订单审核时，验证库存
                bad_skus = _validate_order_stock(order.items, order.warehouse_id)
                # 未通过验证
                if bad_skus:
                    return custom_response(False, gettext("Order[%s] items[%s] isn't exist or Inventory is not enough " % (sn, bad_skus)))

                order.mark_shipment_status()
            else:
                pass

        db.session.commit()
        
    except:
        db.session.rollback()
        return custom_response(False, gettext("Verify order is fail!!!"))

    return full_response(True, R200_OK, selected_sns)


def _validate_order_stock(order_items, warehouse_id):
    """验证订单明细库存"""
    bad_skus = []
    for item in order_items:
        product_stock = ProductStock.query.filter_by(sku_serial_no=item.sku_serial_no,
                                                     warehouse_id=warehouse_id).first()
        # 产品库存不足
        if product_stock is None or product_stock.available_count < item.quantity:
            current_app.logger.warn("Current sku[%s] inventory is not enough!" % item.sku_serial_no)
            bad_skus.append(item.sku_serial_no)
            continue
            
        # 同步减去库存
        product_stock.current_count -= item.quantity
        product_stock.saled_count += item.quantity
        
    return bad_skus
    

@main.route('/orders/ajax_shipped', methods=['POST'])
@login_required
@user_has('admin_order')
def ajax_shipped():
    """订单发货"""
    selected_ids = request.form.getlist('selected[]')
    if not selected_ids or selected_ids is None:
        return custom_response(False, 'Shipped order null!')

    try:
        for rid in selected_ids:
            order = Order.query.filter_by(master_uid=Master.master_uid(), serial_no=rid).first()
            if order and order.status != OrderStatus.SHIPPED:
                order.mark_shipped_status()

        db.session.commit()

    except Exception as ex:
        current_app.logger.warn('Shipped order is fail: %s' % ex)
        return custom_response(False, 'Shipped order is fail: %s' % ex)

    return full_response(True, R200_OK, selected_ids)


@main.route('/orders/<string:rid>/ajax_canceled', methods=['POST'])
@login_required
@user_has('admin_order')
def ajax_canceled(rid):
    """订单取消"""
    order = Order.query.filter_by(serial_no=rid).first()
    if not order:
        return custom_response(False, "Order isn't exist!")

    warehouse_id = order.warehouse_id
    try:
        # 释放订单占用库存
        for item in order.items:
            sku_id = item.sku_id
            product_stock = ProductStock.query.filter_by(product_sku_id=sku_id, warehouse_id=warehouse_id).first()
            if not product_stock:
                raise Exception('Order item stock is exception!')

            product_stock.current_count += 1
            product_stock.saled_count -= 1

        # 订单取消
        order.mark_canceled_status()

        db.session.commit()

    except Exception as e:
        db.session.rollback()
        return custom_response(False, 'Order canceled is fail!')

    return full_response(True, R200_OK, {'rid': rid})


@main.route('/orders/delete', methods=['POST'])
@login_required
@user_has('admin_order')
def delete_order():
    """删除订单"""
    selected_ids = request.form.getlist('selected[]')
    if not selected_ids or selected_ids is None:
        flash(gettext('Delete order null!'), 'danger')
        abort(404)

    try:
        for rid in selected_ids:
            order = Order.query.filter_by(master_uid=Master.master_uid(), serial_no=rid).first()
            if order:
                db.session.delete(order)

        db.session.commit()

    except Exception as ex:
        current_app.logger.debug('Delete order is fail: %s' % ex)

    flash(gettext('Delete order is ok!'), 'success')

    return redirect(url_for('.show_orders'))


@main.route('/orders/download_template')
@login_required
def download_order_tpl():
    dest_filename = r'mic_template_orders.xlsx'
    export_path = current_app.root_path + '/static/tpl/'
    export_file = '{}{}'.format(export_path, dest_filename)

    resp = make_response(send_file(export_file))
    resp.headers['Content-Disposition'] = 'attachment; filename={};'.format(dest_filename)

    return resp
