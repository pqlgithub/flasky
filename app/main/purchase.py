# -*- coding: utf-8 -*-
import datetime, time, hashlib
from jinja2 import PackageLoader, Environment
from flask import render_template, redirect, url_for, abort, flash, request,\
    current_app, make_response, send_file
from flask_sqlalchemy import Pagination
from flask_login import login_required, current_user
from flask_babelex import gettext
from openpyxl.workbook import Workbook
from . import main
from .. import db, uploader
from ..utils import gen_serial_no, full_response, status_response, custom_status, R200_OK, Master, timestamp, custom_response,\
    datestr_to_timestamp, import_product_from_excel
from ..constant import PURCHASE_STATUS, PURCHASE_PAYED, PURCHASE_EXCEL_FIELDS, SORT_TYPE_CODE
from ..decorators import user_has
from app.models import Purchase, PurchaseProduct, Supplier, Product, ProductSku, Warehouse, \
    TransactDetail, InWarehouse, StockHistory, ProductStock, Site
from app.forms import PurchaseForm, PurchaseExpressForm
from .filters import supress_none, timestamp2string, break_line
from ..pdfs import create_pdf

def load_common_data():
    """
    私有方法，装载共用数据
    """
    pending_review_count = Purchase.query.filter_by(master_uid=Master.master_uid(), status=1).count()
    pending_arrival_count = Purchase.query.filter_by(master_uid=Master.master_uid(), status=5).count()
    pending_storage_count = Purchase.query.filter_by(master_uid=Master.master_uid(), status=10).count()
    unpaid_count = Purchase.query.filter_by(master_uid=Master.master_uid(), payed=2).count()
    applyable_count = Purchase.query.filter_by(master_uid=Master.master_uid(), payed=1).count()

    # 库房列表
    warehouse_list = Warehouse.query.filter_by(master_uid=Master.master_uid()).all()

    return {
        'pending_review_count': pending_review_count,
        'pending_arrival_count': pending_arrival_count,
        'pending_storage_count': pending_storage_count,
        'applyable_count': applyable_count,
        'unpaid_count': unpaid_count,
        'purchase_status': PURCHASE_STATUS,
        'purchase_payed': PURCHASE_PAYED,
        'warehouse_list': warehouse_list,
        'top_menu': 'purchases'
    }


@main.route('/purchases')
@main.route('/purchases/<int:page>')
@login_required
@user_has('admin_purchase')
def show_purchases(page=1):
    per_page = request.args.get('per_page', 10, type=int)
    status = request.args.get('s', 0, type=int)
    query = Purchase.query.filter_by(master_uid=Master.master_uid())
    if status:
        query = query.filter_by(status=status)

    paginated_purchases = query.order_by('created_at desc').paginate(page, per_page)

    return render_template('purchases/show_list.html',
                            paginated_purchases=paginated_purchases.items,
                            pagination=paginated_purchases,
                            sub_menu='purchases',
                            status=status,
                            **load_common_data())


@main.route('/purchases/search', methods=['GET', 'POST'])
@login_required
@user_has('admin_purchase')
def search_purchases():
    """支持全文索引搜索采购单"""
    per_page = request.values.get('per_page', 10, type=int)
    page = request.values.get('page', 1, type=int)
    qk = request.values.get('qk')
    wh_id = request.values.get('wh_id', type=int)
    status = request.values.get('s', type=int, default=0)
    days = request.values.get('d', 0, type=int)
    sk = request.values.get('sk', type=str, default='ad')
    ref = request.values.get('ref')

    current_app.logger.debug('qk[%s], sk[%s]' % (qk, sk))

    builder = Purchase.query.filter_by(master_uid=Master.master_uid())
    if qk:
        builder = builder.whoosh_search(qk, like=True)
    if wh_id:
        builder = builder.filter_by(warehouse_id=wh_id)
    if status:
        if ref == 'pay':
            builder = builder.filter_by(payed=status)
        else:
            builder = builder.filter_by(status=status)
    if days:
        offset_days_ago = datetime.date.today() - datetime.timedelta(days)
        builder = builder.filter(Purchase.created_at >= offset_days_ago)

    purchases = builder.order_by('%s desc' % SORT_TYPE_CODE[sk]).all()

    # 构造分页
    total_count = builder.count()
    if page == 1:
        start = 0
    else:
        start = (page - 1) * per_page
    end = start + per_page

    current_app.logger.debug('total count [%d], start [%d], per_page [%d]' % (total_count, start, per_page))

    paginated_purchases = purchases[start:end]
    print(paginated_purchases)
    pagination = Pagination(query=None, page=page, per_page=per_page, total=total_count, items=None)

    tpl = 'purchases/search_pay_result.html' if ref == 'pay' else 'purchases/search_result.html'
    return render_template(tpl,
                           qk=qk,
                           wh_id=wh_id,
                           s=status,
                           d=days,
                           sk=sk,
                           ref=ref,
                           pagination=pagination,
                           paginated_purchases=paginated_purchases)


@main.route('/purchases/payments')
@main.route('/purchases/payments/<int:page>')
@login_required
@user_has('admin_purchase')
def payments(page=1):
    per_page = request.args.get('per_page', 10, type=int)
    status = request.args.get('f', 1, type=int)

    query = Purchase.query.filter_by(master_uid=Master.master_uid(), payed=status)

    paginated_purchases = query.order_by('created_at asc').paginate(page, per_page)

    return render_template('purchases/pay_list.html',
                           paginated_purchases=paginated_purchases.items,
                           pagination=paginated_purchases,
                           sub_menu='purchases',
                           f=status,
                           **load_common_data())


@main.route('/purchases/<string:rid>/preview')
@login_required
@user_has('admin_purchase')
def preview_purchase(rid):
    """预览或查看采购单详情"""
    purchase = Purchase.query.filter_by(serial_no=rid).first()
    # 限制权限
    if purchase is None or purchase.master_uid != Master.master_uid():
        abort(403)

    return render_template('purchases/view_detail.html',
                           purchase=purchase,
                           f=purchase.payed,
                           **load_common_data())


@main.route('/purchases/create', methods=['GET', 'POST'])
@login_required
@user_has('admin_purchase')
def create_purchase():
    form = PurchaseForm()
    if form.validate_on_submit():
        # 获取选中的sku
        sku_list = request.form.getlist('sku[]')
        if not sku_list or sku_list is None:
            flash('Purchase not choose the products!', 'danger')
            return redirect(url_for('.show_purchases'))

        total_quantity = 0
        total_amount = 0
        purchase_products = []
        for sku_id in sku_list:
            sku = {}
            sku_id = int(sku_id)
            sku_row = ProductSku.query.get_or_404(sku_id)

            sku['product_sku_id'] = sku_id
            sku['sku_serial_no'] = sku_row.serial_no
            sku['cost_price'] = float(sku_row.cost_price)
            sku['quantity'] = int(request.form.get('sku[%d][quantity]' % sku_id))

            total_quantity += sku['quantity']
            total_amount += sku['cost_price'] * sku['quantity']

            purchase_products.append(sku)

        purchase = Purchase(
            master_uid=Master.master_uid(),
            serial_no=Purchase.make_unique_serial_no(gen_serial_no('CG')),
            warehouse_id=form.warehouse_id.data,
            supplier_id=form.supplier_id.data,
            freight=form.freight.data,
            extra_charge=form.extra_charge.data,
            arrival_date=form.arrival_date.data,
            description=form.description.data,
            sku_count=len(sku_list),
            total_amount=total_amount,
            quantity_sum=total_quantity
        )
        db.session.add(purchase)

        for purchase_item in purchase_products:
            purchase_item = PurchaseProduct(purchase=purchase, **purchase_item)
            db.session.add(purchase_item)

        db.session.commit()

        flash('Add purchase is success!', 'success')
        return redirect(url_for('.show_purchases'))
    else:
        current_app.logger.debug(form.errors)

    mode = 'create'
    suppliers = Supplier.query.filter_by(master_uid=Master.master_uid()).order_by('created_at desc').all()

    return render_template('purchases/create_and_edit.html',
                           form=form,
                           mode=mode,
                           sub_menu='purchases',
                           purchase=None,
                           suppliers=suppliers,
                           **load_common_data())


@main.route('/purchases/<string:rid>/edit', methods=['GET', 'POST'])
@login_required
@user_has('admin_purchase')
def edit_purchase(rid):
    purchase = Purchase.query.filter_by(serial_no=rid).first()
    # 限制权限
    if purchase is None or purchase.master_uid != Master.master_uid():
        abort(403)

    form = PurchaseForm()
    if form.validate_on_submit():
        # 获取选中的sku
        sku_list = request.form.getlist('sku[]')
        if not sku_list or sku_list is None:
            flash('Purchase not choose the products!', 'danger')
            return redirect(url_for('.show_purchases'))

        # 清空旧产品
        for old_item in purchase.products:
            db.session.delete(old_item)

        total_quantity = 0
        total_amount = 0
        for sku_id in sku_list:
            sku = {}
            sku_id = int(sku_id)
            sku_row = ProductSku.query.get_or_404(sku_id)

            sku['product_sku_id'] = sku_id
            sku['sku_serial_no'] = sku_row.serial_no
            sku['cost_price'] = float(sku_row.cost_price)
            sku['quantity'] = int(request.form.get('sku[%d][quantity]' % sku_id))

            total_quantity += sku['quantity']
            total_amount += sku['cost_price']*sku['quantity']

            # 添加新产品
            purchase_item = PurchaseProduct(purchase=purchase, **sku)
            db.session.add(purchase_item)

        purchase.warehouse_id = form.warehouse_id.data,
        purchase.supplier_id = form.supplier_id.data,
        purchase.freight = form.freight.data,
        purchase.extra_charge = form.extra_charge.data,
        purchase.arrival_date = form.arrival_date.data,
        purchase.description = form.description.data
        purchase.quantity_sum = total_quantity
        purchase.total_amount = total_amount
        purchase.sku_count = len(sku_list)

        db.session.commit()

        flash('Edit purchase is success!', 'success')
        return redirect(url_for('.show_purchases'))
    else:
        current_app.logger.debug(form.errors)

    mode = 'edit'
    suppliers = Supplier.query.filter_by(master_uid=Master.master_uid()).order_by('created_at desc').all()

    form.warehouse_id.data = purchase.warehouse_id
    form.supplier_id.data = purchase.supplier_id
    form.freight.data = purchase.freight
    form.extra_charge.data = purchase.extra_charge
    form.arrival_date.data = purchase.arrival_date
    form.description.data = purchase.description

    return render_template('purchases/create_and_edit.html',
                           form=form,
                           mode=mode,
                           sub_menu='purchases',
                           purchase=purchase,
                           suppliers=suppliers,
                           **load_common_data())


@main.route('/purchases/delete', methods=['POST'])
@login_required
@user_has('admin_purchase')
def delete_purchase():
    selected_ids = request.form.getlist('selected[]')
    if not selected_ids or selected_ids is None:
        flash('Delete purchase is null!', 'danger')
        abort(404)

    try:
        for rid in selected_ids:
            purchase = Purchase.query.filter_by(serial_no=rid).first()
            # 验证权限
            if purchase.master_uid != Master.master_uid():
                continue
            db.session.delete(purchase)
        db.session.commit()

        flash('Delete purchase is ok!', 'success')
    except:
        db.session.rollback()
        flash('Delete purchase is fail!', 'danger')

    return redirect(url_for('.show_purchases'))


@main.route('/purchases/delete_item', methods=['POST'])
@login_required
@user_has('admin_purchase')
def delete_purchase_item():
    item_id = request.form.get('item_id')
    purchase_item = PurchaseProduct.query.get(int(item_id))
    if not purchase_item:
        return custom_status('Purchase item is not exist!')

    sku_id = purchase_item.product_sku_id
    purchase_id = purchase_item.purchase_id
    purchase = Purchase.query.get(purchase_id)
    if not purchase:
        return custom_status('Purchase is not exist!')

    # 更新采购单数量、总金额
    purchase.sku_count -= 1
    purchase.quantity_sum -= purchase_item.quantity
    purchase.total_amount -= purchase_item.quantity*purchase_item.cost_price

    db.session.delete(purchase_item)
    db.session.commit()

    return full_response(True, R200_OK, {'sku_id': sku_id})


@main.route('/purchases/ajax_verify', methods=['POST'])
@login_required
@user_has('admin_purchase')
def ajax_verify():
    selected_ids = request.form.getlist('selected[]')
    if not selected_ids or selected_ids is None:
        return custom_response(False, gettext('Verify purchase id is NULL!'))

    try:
        for id in selected_ids:
            purchase = Purchase.query.filter_by(serial_no=id).first()
            if purchase is None:
                return custom_response(False, gettext('Verify purchase is Null!'))
            if purchase.master_uid != Master.master_uid():
                return custom_response(False, gettext('You do not have permission to operate'))

            # 审批完毕，待到货状态
            purchase.update_status(5)

        db.session.commit()
    except:
        db.session.rollback()
        return status_response(False, custom_status('Verify purchase is fail!'))

    return full_response(True, R200_OK, selected_ids)


@main.route('/purchases/<string:rid>/ajax_arrival', methods=['GET', 'POST'])
@login_required
@user_has('admin_purchase')
def ajax_arrival(rid):
    purchase = Purchase.query.filter_by(serial_no=rid).first()
    if request.method == 'POST':
        quantity_offset = 0
        detail_products = []
        remark = request.form.get('remark')
        items = request.form.getlist('items[]')
        for sku_id in items:
            quantity = request.form.get('quantity[%s]' % sku_id, 0, type=int)
            if not quantity:
                continue

            item_product = purchase.products.filter_by(product_sku_id=sku_id).first()
            if not item_product.validate_quantity(quantity):
                return status_response(False, custom_status('Quantity exceeded total quantity!'))

            # 更新入库数量
            item_product.in_quantity += quantity
            # 新增总数量
            quantity_offset += quantity

            # 入库单明细
            ori_stock = ProductStock.get_stock_quantity(purchase.warehouse_id, sku_id)
            stock_history = StockHistory(
                master_uid=Master.master_uid(),
                warehouse_id=purchase.warehouse_id,
                product_sku_id=sku_id,
                sku_serial_no=item_product.sku_serial_no,
                type=1,
                operation_type=10,
                original_quantity=ori_stock,
                quantity=quantity,
                current_quantity=ori_stock+quantity,
                ori_price=item_product.cost_price,
                price=item_product.cost_price,
            )
            detail_products.append(stock_history)

            # 更新库房累计库存数
            product_stock = ProductStock.validate_is_exist(purchase.warehouse_id, sku_id)
            if not product_stock:
                new_stock = ProductStock(
                    master_uid=Master.master_uid(),
                    product_sku_id=sku_id,
                    sku_serial_no=item_product.sku_serial_no,
                    warehouse_id=purchase.warehouse_id,
                    total_count=quantity,
                    current_count=quantity
                )
                db.session.add(new_stock)
            else:
                product_stock.total_count += quantity
                product_stock.current_count += quantity

        # 自动生成入库单
        in_serial_no = gen_serial_no('RK')
        in_warehouse = InWarehouse(
            master_uid=purchase.master_uid,
            serial_no=in_serial_no,
            target_serial_no=purchase.serial_no,
            warehouse_id=purchase.warehouse_id,
            total_quantity=quantity_offset,
            remark=remark
        )
        db.session.add(in_warehouse)
        # 添加操作明细
        for sh in detail_products:
            sh.serial_no = in_serial_no
            db.session.add(sh)

        # 检测是否完成入库
        if purchase.validate_finished(quantity_offset):
            # 入库完成
            purchase.update_status(15)
            purchase.arrival_at = timestamp()
        purchase.in_quantity += quantity_offset

        db.session.commit()

        return status_response(True, R200_OK)

    return render_template('purchases/_modal_arrival.html',
                           purchase=purchase,
                           post_url=url_for('.ajax_arrival', rid=rid))


@main.route('/purchases/ajax_apply_pay', methods=['POST'])
@login_required
@user_has('admin_purchase')
def ajax_apply_pay():
    selected_ids = request.form.getlist('selected[]')
    if not selected_ids or selected_ids is None:
        return status_response(False, custom_status('Apply purchase pay, id is NULL!'))

    try:
        for rid in selected_ids:
            purchase = Purchase.query.filter_by(serial_no=rid).first()

            if purchase is None:
                return status_response(False, custom_status('Apply purchase pay, id[%s] is NULL!' % rid))
            if purchase.payed != 1:
                return status_response(False, custom_status('Purchase[%s] is already payed!' % rid))
            if purchase.master_uid != Master.master_uid():
                return custom_response(False, gettext('You do not have permission to operate'))

            # 申请付款，进入待付款流程
            purchase.update_payed(2)

            # 自动同步产生待付款单
            transaction = TransactDetail(
                master_uid=Master.master_uid(),
                serial_no=gen_serial_no('FK'),
                type=2,
                transact_user=purchase.supplier.short_name,
                amount=purchase.total_amount,
                target_id=purchase.id,
                target_type=1
            )
            db.session.add(transaction)

        db.session.commit()
    except Exception as ex:
        db.session.rollback()
        current_app.logger.debug('Apply purchase is fail [%s]!' % ex)
        return status_response(False, custom_status('Apply purchase is fail!'))

    return full_response(True, R200_OK, selected_ids)


@main.route('/purchases/<string:rid>/add_express_no', methods=['GET', 'POST'])
@login_required
@user_has('admin_purchase')
def purchase_express_no(rid):
    purchase = Purchase.query.filter_by(serial_no=rid).first()
    if purchase is None or purchase.master_uid != Master.master_uid():
        return gettext('You do not have permission to operate')

    form = PurchaseExpressForm()
    if form.validate_on_submit():
        purchase.express_name = form.express_name.data
        purchase.express_no = form.express_no.data

        db.session.commit()

        return status_response(True, R200_OK)

    form.express_name.data = purchase.express_name
    form.express_no.data = purchase.express_no
    return render_template('purchases/_modal_express.html',
                           purchase=purchase,
                           form=form,
                           post_url=url_for('.purchase_express_no', rid=rid))


@main.route('/purchases/output_purchase')
@login_required
@user_has('admin_purchase')
def output_purchase():
    rid = request.args.get('rid')
    rids = rid.split(',')
    purchase_list = Purchase.query.filter(Purchase.serial_no.in_(rids)).all()
    return render_template('pdf/purchase.html',
                           purchase_list=purchase_list)


@main.route('/purchases/print_purchase_pdf')
@login_required
@user_has('admin_purchase')
def print_purchase_pdf():
    """输出pdf，并打印"""
    rid = request.args.get('rid')
    rids = rid.split(',')
    purchase_list = Purchase.query.filter(Purchase.serial_no.in_(rids)).all()

    env = Environment(loader=PackageLoader(current_app.name, 'templates'))
    env.filters['supress_none'] = supress_none
    env.filters['timestamp2string'] = timestamp2string
    env.filters['break_line'] = break_line
    template = env.get_template('pdf/purchase.html')

    current_site = Site.query.filter_by(master_uid=Master.master_uid()).first()

    title_attrs = {
        'serial_no': gettext('Purchase Serial'),
        'status': gettext('Status'),
        'supplier' : gettext('Supplier'),
        'supplier_info': gettext('Supplier Info'),
        'warehouse_info': gettext('Warehouse Info'),
        'warehouse_name' : gettext('Warehouse Name'),
        'date' : gettext('Date'),
        'contact_name' : gettext('Contact name'),
        'address': gettext('Address'),
        'phone': gettext('Phone'),
        'email': gettext('E-mail'),
        'remark': gettext('Remark'),
        'sn': gettext('Serial Number'),
        'product_info': gettext('Product Info'),
        'price': gettext('Price'),
        'quantity': gettext('Purchase Quantity'),
        'in_quantity': gettext('Arrival Quantity'),
        'subtotal': gettext('Subtotal'),
        'express_no': gettext('Express No.'),
        'freight': gettext('Freight'),
        'other_charge': gettext('Other Charge'),
        'total_amount': gettext('Total Amount'),
        'arrival_quantity': gettext('Arrival Quantity')
    }

    # code_bars = {}
    # options = dict(text_distance=2,font_size=16)
    # root_path = current_app.root_path + '/static/code_bars/'
    # for sn in rids:
    #    ean = barcode.get('code39', sn, writer=ImageWriter())
    #    filename = 'serial_no_' + sn
    #    code_bars[sn] = ean.save(root_path + filename, options)

    html = template.render(
        current_site=current_site,
        title_attrs=title_attrs,
        font_path=current_app.root_path + '/static/fonts/simsun.ttf',
        purchase_list=purchase_list,
    ).encode('utf-8')

    pdf = create_pdf(html)

    current_app.logger.debug('Pdf create ok!')

    resp = make_response(pdf.getvalue())

    current_app.logger.debug('Make response is ok!')

    export_file = 'Purchase-{}'.format(int(timestamp()))
    resp.headers['Content-Disposition'] = ("inline; filename='{0}'; filename*=UTF-8''{0}".format(export_file))
    resp.headers['Content-Type'] = 'application/pdf'
    
    return resp


@main.route('/purchases/export_purchase', methods=['GET', 'POST'])
@login_required
@user_has('admin_purchase')
def export_purchase():
    per_page = request.args.get('per_page', 30, type=int)
    if request.method == 'POST':
        wh_id = request.form.get('wh_id', 0, type=int)
        status = request.form.getlist('status[]')
        start_date = request.form.get('start_date')
        end_date = request.form.get('end_date')

        builder = Purchase.query.filter_by(master_uid=Master.master_uid())
        if wh_id:
            builder = builder.filter_by(warehouse_id=wh_id)
        if status and len(status):
            status = [int(s) for s in status]
            builder = builder.filter(Purchase.status.in_(status))

        if start_date:
            builder = builder.filter(Purchase.created_at >= datestr_to_timestamp(start_date))

        if end_date:
            builder = builder.filter(Purchase.created_at <= datestr_to_timestamp(end_date))


        purchase_list = builder.order_by('created_at desc').all()

        dest_filename = r'mic_purchase_list_{}.xlsx'.format(datetime.datetime.now().strftime('%Y%m%d'))
        export_path = current_app.root_path + '/static/'

        # 新建文件
        wb = Workbook()

        # 第一个sheet
        ws = wb.active
        ws.title = 'Purchase List'

        current_app.logger.debug('Purchase count: %d' % len(purchase_list))

        purchase_columns = [key for key in PURCHASE_EXCEL_FIELDS.keys()]
        current_site = Site.query.filter_by(master_uid=Master.master_uid()).first()

        # 写入表头
        for col in range(1, len(purchase_columns) + 1):
            field = purchase_columns[col - 1]
            ws.cell(row=1, column=col, value=_rebuild_header_value(field, current_site))

        # 写入数据, 从第2行开始写入
        for row in range(2, len(purchase_list) + 2):
            purchase = purchase_list[row - 2]
            warehouse = purchase.warehouse
            supplier = purchase.supplier

            current_app.logger.debug('Current row: %d' % row)

            for col in range(1, len(purchase_columns) + 1):
                field = purchase_columns[col - 1]

                current_app.logger.debug('Current col: %s' % field)

                if field == 'warehouse':
                    cell_value = warehouse.to_json().get('name')
                elif field == 'supplier':
                    cell_value = supplier.to_json().get('short_name')
                elif field in ['contact_name', 'phone']:
                    cell_value = supplier.to_json().get(field)
                elif field in ['created_at', 'arrival_at']:
                    cell_value = timestamp2string(purchase.to_json().get(field)) if purchase.to_json().get(field) > 0 else ''
                elif field == 'product_name':
                    cell_value = purchase.product_name
                elif field == 'product_sku':
                    cell_value = purchase.product_sku
                elif field == 'status':
                    cell_value = purchase.status_label[1]
                else:
                    cell_value = purchase.to_json().get(field)

                ws.cell(row=row, column=col, value=cell_value)

        export_file = '{}{}'.format(export_path, dest_filename)
        wb.save(filename=export_file)

        resp = make_response(send_file(export_file))
        resp.headers['Content-Disposition'] = 'attachment; filename={};'.format(dest_filename)

        return resp

    # 库房列表
    warehouse_list = Warehouse.query.filter_by(master_uid=Master.master_uid()).all()

    return render_template('purchases/_modal_export.html',
                           warehouse_list=warehouse_list,
                           purchase_status=PURCHASE_STATUS)


def _rebuild_header_value(key, current_site):
    """重建Excel表头标题"""
    if key in ['freight','extra_charge','total_amount',]:
        return '{}({})'.format(PURCHASE_EXCEL_FIELDS[key], current_site.currency)
    return PURCHASE_EXCEL_FIELDS[key]


@main.route('/purchases/import_purchase', methods=['GET', 'POST'])
@login_required
@user_has('admin_purchase')
def import_purchase():
    if request.method == 'POST':
        wh_id = request.form.get('wh_id', type=int)
        f = request.files['excel']

        # start to save
        sub_folder = str(time.strftime('%y%m%d'))
        name_prefix = 'admin' + str(time.time())
        name_prefix = hashlib.md5(name_prefix.encode('utf-8')).hexdigest()[:15]
        filename = uploader.save(f, folder=sub_folder, name=name_prefix + '.')

        storage_filepath = uploader.path(filename)
        current_app.logger.debug('Excel file [%s]' % storage_filepath)

        # 读取文档内容
        product_skus = import_product_from_excel(storage_filepath)

        total_quantity = 0
        total_amount = 0
        total_count = 0
        purchase_products = []
        supplier_id = 0
        # fields = ['name', 'mode', 'color', 'id_code', 'cost_price', 'quantity']

        for sku_dict in product_skus:
            sku_id_code = sku_dict.get('id_code')
            # 无69码，跳过
            if sku_id_code is None:
                continue

            # 通过69码查找匹配的产品
            sku_row = ProductSku.query.filter_by(master_uid=Master.master_uid(), id_code=sku_id_code).first()

            if sku_row is None:
                continue

            sku = {}
            sku['product_sku_id'] = sku_row.id
            sku['sku_serial_no'] = sku_row.serial_no
            sku['cost_price'] = float(sku_dict.get('cost_price'))
            sku['quantity'] = int(sku_dict.get('quantity', 0))

            supplier_id = sku_row.supplier_id

            total_quantity += sku['quantity']
            total_amount += sku['cost_price'] * sku['quantity']
            total_count += 1

            purchase_products.append(sku)

        purchase = Purchase(
            master_uid=Master.master_uid(),
            serial_no=Purchase.make_unique_serial_no(gen_serial_no('CG')),
            warehouse_id=wh_id,
            supplier_id=supplier_id,
            sku_count=total_count,
            total_amount=total_amount,
            quantity_sum=total_quantity
        )

        db.session.add(purchase)

        for purchase_item in purchase_products:
            purchase_item = PurchaseProduct(purchase=purchase, **purchase_item)
            db.session.add(purchase_item)

        db.session.commit()

        return redirect(url_for('.show_purchases'))

    # 库房列表
    warehouse_list = Warehouse.query.filter_by(master_uid=Master.master_uid()).all()
    paginated_suppliers = Supplier.query.filter_by(master_uid=Master.master_uid()).order_by('created_at desc').paginate(
        1, 1000)
    return render_template('purchases/_modal_import.html',
                           warehouse_list=warehouse_list,
                           paginated_suppliers=paginated_suppliers)


@main.route('/purchases/download_template')
@login_required
def download_purchase_tpl():
    dest_filename = r'mic_template_purchasing.xlsx'
    export_path = current_app.root_path + '/static/tpl/'
    export_file = '{}{}'.format(export_path, dest_filename)

    resp = make_response(send_file(export_file))
    resp.headers['Content-Disposition'] = 'attachment; filename={};'.format(dest_filename)

    return resp