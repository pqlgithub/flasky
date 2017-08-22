# -*- coding: utf-8 -*-
import sys, os, ssl, time
from urllib import request, parse, error
from bs4 import BeautifulSoup
from openpyxl.reader.excel import load_workbook
from openpyxl.workbook import Workbook
from flask_script import Command

basedir = os.path.abspath(os.path.dirname(os.path.dirname(__file__)))

file_path = '%s/public/data/amazon.xlsx' % basedir

class AmazonSearcher(Command):

    def run(self):
        if not os.path.exists(file_path):
            print("File is't exist!")
            sys.exit(0)

        # 读取文件
        wb = load_workbook(filename=file_path)
        sheets = wb.get_sheet_names()
        # 默认第一个表格的名称
        ws = wb.get_sheet_by_name(sheets[0])

        # 获取行数
        total_rows = ws.max_row

        for row_idx in xrange(7, total_rows + 1):
            cell_value = ws.cell(row=row_idx, column=2).value
            if cell_value:
                time.sleep(15)
                count_text = self.get_page_html(cell_value)
                ws.cell(row=row_idx, column=3, value=count_text)

        wb.save(file_path)
        
        sys.exit(0)


    def get_page_html(self, word):
        word = parse.quote_plus(word)
        url = 'https://www.amazon.com/s/ref=nb_sb_noss_2?url=search-alias%3Daps&field-keywords={}'.format(word)

        headers = {
            'User-Agent': r'Mozilla/5.0 (Windows NT 6.1; WOW64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/45.0.2454.85 Safari/537.36 115Browser/6.0.3',
            'Referer': r'http://www.baidu.com',
            'Connection': 'keep-alive'
        }

        try:
            ssl._create_default_https_context = ssl._create_unverified_context

            print('Start to request [%s]' % url)

            req = request.Request(url)
            response = request.urlopen(req).read()
            response = response.decode('utf-8')

            print('Request is ok.')

            soup = BeautifulSoup(response, 'html.parser')

            h2 = soup.find_all(id="s-result-count")

            count_text = h2[0].get_text()

            return count_text.split(' ')[2]

        except error.HTTPError as ex:
            print(ex.read().decode('utf-8'))

        return 0